from os import listdir
import shutil
from docx import Document
import pandas as pd
from openpyxl import load_workbook, workbook
import telebot
from telebot import types
import time
from random import randint
from configparser import ConfigParser
from os import path
import os


def checkValidateId(id):
    q = ConfigParser()
    q.read("adminId.ini")
    if not path.exists("adminId.ini"):
        with open('adminId.ini', 'wb') as configfile:
            q.write(configfile)
    if str(id) in q.sections():
        return True
    else:
        return False


def addAdminId(id):
    if not str(id).isdigit():
        return False
    q = ConfigParser()
    q.clear()
    q.read("adminId.ini")
    if not path.exists("adminId.ini"):
        with open('adminId.ini', 'wb') as configfile:
            q.write(configfile)
    try:
        q.add_section(str(id))
        with open('adminId.ini', 'w') as configfile:
            q.write(configfile)
        return True
    except:
        return True


def checkValidateHash(hash):
    q = ConfigParser()
    q.read("adminHash.ini")
    if not path.exists("adminHash.ini"):
        with open('adminHash.ini', 'wb') as configfile:
            q.write(configfile)
    a = q.sections()
    if hash in a:
        q.remove_section(hash)
        with open('adminHash.ini', 'w') as configfile:
            q.write(configfile)
        return True
    else:
        return False


def lastMsgGet(id):
    id = str(id)
    q = ConfigParser()
    q.clear()
    q.read("lastMsg.ini")
    if not path.exists("lastMsg.ini"):
        with open('lastMsg.ini', 'wb') as configfile:
            q.write(configfile)
    try:
        q.clear()
        q.read("lastMsg.ini")
        if id in q.sections():
            ret = q.get(id, "sender")
            q.remove_section(id)
            with open('lastMsg.ini', 'w') as configfile:
                q.write(configfile)
            return ret
        else:
            return False
    except Exception as e:
        print(str(e))
        return False

def lastMsgAdd(id, mode):
    q = ConfigParser()
    q.clear()
    q.read("lastMsg.ini")
    if not path.exists("lastMsg.ini"):
        with open('lastMsg.ini', 'wb') as configfile:
            q.write(configfile)
    try:
        q.clear()
        q.read("lastMsg.ini")
        id = str(id)
        q.add_section(id)
        if mode=="1":
            q.set(id, "sender", "1")
        elif mode=='2':
            q.set(id, "sender", "2")
        elif mode=='3':
            q.set(id, "sender", "3")
        with open('lastMsg.ini', 'w') as configfile:
            q.write(configfile)
        return True
    except Exception as e:
        print(str(e))
        return False

def randomHash():
    q = "1234567890"
    rand = ""

    try:
        for i in range(5):
            rand += q[randint(0, len(q)-1)]
        q = ConfigParser()
        q.read("adminHash.ini")
        if not path.exists("adminHash.ini"):
            with open('adminHash.ini', 'wb') as configfile:
                q.write(configfile)
        q.clear()
        q.read("adminHash.ini")
        q.add_section(rand)
        with open("adminHash.ini", "w") as cf:
            q.write(cf)
    except Exception as e:
        print(str(e))
    return rand


def checkScheudleExists():
    if not path.exists("floder"):
        return False
    fileArray = listdir(path="floder/")
    result = False
    for i in fileArray:
        if i[-5:] == ".docx":
            result = i
    return result


def scheudleDate():
    q = os.listdir("floder/")
    for i in q:
        if i[-5:] == ".docx":
            return i[:9]


def getscheudle(group):
    ret = []
    wb = load_workbook("floder/Table# 0.xlsx")
    wb = wb.active
    wb1 = load_workbook("floder/Table# 1.xlsx")
    wb1 = wb1.active

    group = group

    limit = 200
    baseRow = 2
    row = 2
    column = 1

    while True:
            valueOfCell = wb.cell(row=row, column=column).value
            if valueOfCell == None and row == limit and column > 10:
                break
            elif valueOfCell == None and row == limit:
                row = baseRow
                column += 1
            else:                                     
                if valueOfCell == group or group in str(valueOfCell):
                    q1 = wb.cell(row=row, column=2).value + " " + wb.cell(row=row, column=column-1).value
                    ret.append(q1)
                    row = baseRow
                    column += 1
                else:
                    row += 1
    if len(ret) == 0:
        limit = 200
        baseRow = 2
        row = 2
        column = 1
        while True:
            valueOfCell = wb1.cell(row=row, column=column).value
            if valueOfCell == None and row == limit and column > 10:
                break
            elif valueOfCell == None and row == limit:
                row = baseRow
                column += 1
            else:                       
                if valueOfCell == group or group in str(valueOfCell):
                    q1 = str(wb1.cell(row=row, column=2).value) + " " + str(wb1.cell(row=row, column=column-1).value)
                    ret.append(q1)     
                    row = baseRow
                    column += 1
                else:
                    row += 1
    return ret


def newUserGet(id):
    q = ConfigParser()
    q.clear()
    q.read("usersId.ini")
    if not path.exists("usersId.ini"):
        with open('usersId.ini', 'wb') as configfile:
            q.write(configfile)
    q.add_section(str(id))
    with open('usersId.ini', 'w') as configfile:
            q.write(configfile)


bot=telebot.TeleBot("5928010197:AAEhljliZrigpF8JBCNTvrSmQrcP_GzMAlQ")


@bot.message_handler(content_types=['text'])
def start_message(message):
    data = lastMsgGet(id=message.chat.id)
    if data:
        if message.text == "/cancel":
            bot.send_message(message.chat.id, text="Отмененно")
        elif data == "1":
            q = ConfigParser()
            q.clear()
            q.read("usersId.ini")
            for i in q.sections():
                bot.send_message(int(i), text=message.text)
        elif data == "2":
            if addAdminId(id=message.text):
                bot.send_message(message.chat.id, text="Админ успешно добавлен")
            else:
                bot.send_message(message.chat.id, text="Не удалось добавить админа")
        elif data == "3":
            q = ConfigParser()
            q.clear()
            q.read("adminId.ini")
            for i in q.sections():
                msg = ""
                msg = f"Сообщение от {message.from_user.first_name} {message.from_user.last_name}, @{message.from_user.username}, \n\n'{message.text}'"
                bot.send_message(int(i), text=msg)
            bot.send_message(message.chat.id, text="Ваше сообщение отправленно администрации канала.\nСпасибо что помогаете улучшать бота")
    elif message.text == "Сгенерировать ключ использования" and checkValidateId(message.chat.id):
        q = randomHash()
        bot.send_message(message.chat.id, text=f"Вы получили одноразовый админ ключ.\nКлюч: {q}")
    elif message.text == "Сделать рассылку" and checkValidateId(id=message.chat.id):
        lastMsgAdd(id=message.chat.id, mode="1")
        bot.send_message(message.chat.id, text="Отправьте сообщение которое хотите разослать всем пользователям.\n'/cancel'\nДля отмены")
    elif message.text == "Админ панель":
        if checkValidateId(message.chat.id):
            markup = types.ReplyKeyboardMarkup()
            markup1 = types.KeyboardButton(text="Получить файл с расписанием")
            markup2 = types.KeyboardButton(text="Добавить админа по ид")
            markup3 = types.KeyboardButton(text="Сгенерировать ключ использования")
            markup4 = types.KeyboardButton(text="Сделать рассылку")
            markup.add(markup1)
            markup.add(markup2)
            markup.add(markup3)
            markup.add(markup4)
            bot.send_message(message.chat.id, text=f"Вы вошли в админ панель, вы можетеиспользовать кнокпи для администрирования ботом", reply_markup=markup)
        else:
            markup = types.ReplyKeyboardMarkup()
            markup.add(types.KeyboardButton(request_contact=True, text="Отправить данные"))
            bot.send_message(message.chat.id, text="Отправьте данные вашего аккаунта для верификации", reply_markup=markup)
    elif "Добавить админа по ид" in message.text:
            if (checkValidateId(id=message.chat.id)):
                lastMsgAdd(id=message.chat.id, mode="2")
                bot.send_message(message.chat.id, text="Отправьте id админа для добовления, или напишите\n'/cancel'\nдля отмены")
            else:
                try:
                    text = str(message.text).replace("Добавить админа по ид ", "")
                    if checkValidateHash(text):
                        lastMsgAdd(id=message.chat.id, mode="2")
                        bot.send_message(message.chat.id, text="Отправьте id админа для добовления, или напишите\n'/cancel'\nдля отмены")
                    else:
                        raise
                except:
                    bot.send_message(message.chat.id, text="Ошибка при верификации")
    elif message.text == "Запросить файл с расписанием":
        data = checkScheudleExists()
        if not data:
            bot.send_message(message.chat.id, text="Не удалось найти файл с расписанием")
        else:
            f = open("floder/" + data, "rb")
            bot.send_document(message.chat.id,f)
    elif message.text == "/start":
        try:
            newUserGet(message.chat.id)
        except:
            pass
        markup = types.ReplyKeyboardMarkup()
        markup.add(types.KeyboardButton(text="Нашли баг, или есть предложение по улучшению бота."))
        bot.send_message(message.chat.id, text="Здраствуйте!\nНапишите название вашей группы и получите последнее расписание!", reply_markup=markup)
    elif message.text == "Нашли баг, или есть предложение по улучшению бота.":
        lastMsgAdd(id=message.chat.id, mode="3")
        bot.send_message(message.chat.id,text="Опишите текстовым сообщение что именно вы хотели бы донести до администраци.\n'/cancel'\nДля отмены")
    elif message.text == "/getid":
        bot.send_message(message.chat.id, text=f"id: {message.chat.id}")
    else:
        try:
            b = getscheudle(group=str(message.text).upper())
            if len(b) == 0:
                raise Exception
            q = ""
            a = 1
            for i in b:
                if a == 4:
                    q += f"{a} Пара: " + str(i) + " (Онлайн)" + "\n"
                else:
                    i = str(i).replace("\n", " ")
                    q += f"{a} Пара: " + str(i) + "\n"
                a += 1
            bot.send_message(message.chat.id, text="Расписание для группы " + str(message.text).upper() + " на " + scheudleDate() + "\n" + q)
        except Exception as e:
            print(str(e))
            bot.send_message(message.chat.id, text="Не удалось найти расписания для данной группы")


@bot.message_handler(content_types=['contact'])
def contact(message):
    if message.contact is not None and message.contact.phone_number == "77087536109":
        markup = types.ReplyKeyboardMarkup()
        markup1 = types.KeyboardButton(text="Запросить файл с расписанием")
        markup2 = types.KeyboardButton(text="Добавить админа по ид")
        markup3 = types.KeyboardButton(text="Сгенерировать ключ использования")
        markup4 = types.KeyboardButton(text="Сделать рассылку")
        markup.add(markup1)
        markup.add(markup2)
        markup.add(markup3)
        markup.add(markup4)
        addAdminId(message.chat.id)
        bot.send_message(message.chat.id, text=f"Вы вошли в админ панель, вы можетеиспользовать кнокпи для администрирования ботом, ваш ид {message.chat.id}", reply_markup=markup)
    else:
        bot.send_message(message.chat.id, text="Не удалось войти в админ панель")

@bot.message_handler(content_types=['document'])
def handle_docs_photo(message):
    try:
        if checkValidateHash(hash=message.caption) or checkValidateId(message.chat.id):
            file_info = bot.get_file(message.document.file_id)
            downloaded_file = bot.download_file(file_info.file_path)
            if os.path.exists("floder"):
                shutil.rmtree("floder")
            os.makedirs("floder")
            src = message.document.file_name
            with open(src, 'wb') as new_file:
                new_file.write(downloaded_file)
            shutil.move(message.document.file_name, "floder/"+message.document.file_name)
            document = Document("floder/"+message.document.file_name)
            for index,table in enumerate(document.tables):
                df = [['' for i in range(len(table.columns))] for j in range(len(table.rows))]
                for i, row in enumerate(table.rows):
                        for j, cell in enumerate(row.cells):
                            df[i][j] = cell.text
                        pd.DataFrame(df).to_excel("Table# "+str(index)+".xlsx")
                        shutil.move("Table# "+str(index)+".xlsx", "floder/Table# "+str(index)+".xlsx")
            bot.reply_to(message, "Файл с расписание успешно принят!")
        else:
            bot.reply_to(message, "Не удалось добавить файл")
    except Exception as e:
        print(str(e))
        bot.reply_to(message, text="Извините но произошла ошибка!")
bot.polling(non_stop=True)

