from datetime import datetime
from os import listdir
import shutil
from docx import Document
from openpyxl import load_workbook, Workbook
import telebot
import os

bot=telebot.TeleBot("5928010197:AAEhljliZrigpF8JBCNTvrSmQrcP_GzMAlQ")

@bot.message_handler(content_types=['document'])
def handle_docs_photo(message):
    try:
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        src = message.document.file_name
        srcEight = src[:6] + src[8:10]
        current_date = datetime.now()
        formatted_date = current_date.strftime("%d.%m.%y")
        pathDate = f"package/{formatted_date}/"
        if srcEight == formatted_date:
            pathDate = f"package/{formatted_date}/"
            xlsx_path = f'{pathDate}{formatted_date}.xlsx'
            if not os.path.exists(pathDate):
                os.makedirs(pathDate)
            with open(pathDate + src, 'wb') as new_file:
                new_file.write(downloaded_file)
        else:
            pathDate = f"package/{srcEight}/"
            xlsx_path = f'{pathDate}{srcEight}.xlsx'
            if not os.path.exists(pathDate):
                os.makedirs(pathDate)
            with open(pathDate + src, 'wb') as new_file:
                new_file.write(downloaded_file)
        docx_path = pathDate + src
        doc = Document(docx_path)
        tables = []
        for table in doc.tables:
            table_data = []
            for row in table.rows:
                row_data = []
                for cell in row.cells:
                    row_data.append(cell.text)
                table_data.append(row_data)
            tables.append(table_data)
        
        wb = Workbook()
        wb.remove(wb.active)
        for idx, table in enumerate(tables):
            ws = wb.create_sheet(title=f'Table {idx + 1}')
            for row in table:
                ws.append(row)
        wb.save(xlsx_path)

        bot.reply_to(message, "Файл с расписание успешно принят!")
    except Exception as e:
        print(str(e))
        bot.reply_to(message, text="Произошла ошибка!")
bot.polling(non_stop=True)

