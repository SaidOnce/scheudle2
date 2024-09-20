from openpyxl import load_workbook
from flask import Flask, jsonify, request
from flask_cors import CORS
from datetime import datetime, timedelta
from os import path
from dbControl import DB


app = Flask(__name__)
CORS(app)


@app.route("/api", methods=['POST'])
def api():
    data = request.json
    username = data.get("username")
    mail = data.get("mail")
    password = data.get("password")
    repeatPassword = data.get("repeatPassword")

    if repeatPassword == password:
        DB().createUser(userName=username, mail=mail, password=password)
        response = {"message": 200, "receivedData": "successfly!"}
        return jsonify(response)
    response = {"message": 300, "receivedData": "Error not the same password and rPassword"}
    return jsonify(response)

@app.route("/apiCheck", methods=['POST'])
def apiCheck():
    data = request.json
    username = data.get("username")
    password = data.get("password")
    if DB().checkValidateUser(userName=username, password=password):
        response = {"message": 200}
        return jsonify(response)
    response = {"message": 300}
    return jsonify(response)



@app.route("/removeUser", methods=['POST'])
def removeUser():
    data = request.json
    key = data.get("key")
    DB().removeUser(key=key)
    response = {"message": 200}
    return jsonify(response)


@app.route("/userList", methods=['GET'])
def userList():
    response = {"list": DB().getUserList()}
    return jsonify(response)


@app.route('/data', methods=['GET', "POST"])
def get_data():
    if request.method == "POST":
        return jsonify(200)
    if request.method == "GET":
        array1 = [["Пусто"],["Пусто"],["Пусто"],["Пусто"],["Пусто"],["Пусто"],["Пусто"]]
        current_date = datetime.now()
        current_weekday = current_date.weekday()
        days_since_monday = current_weekday
        monday_date = current_date - timedelta(days=days_since_monday)
        dates_of_current_week = [monday_date + timedelta(days=i) for i in range(7)]
        for index, i in enumerate(dates_of_current_week):
            i = i.strftime('%d.%m.%y')
            array1[index] = [i]
            print(i)
            try:
                scheudle = getscheudle(group="Т22-3Б", date=f"package/{i}/{i}.xlsx")
                scheudle = [i] + scheudle
                if scheudle:
                    array1[index] = scheudle
            except:
                pass
        data = {
           "s" : array1,
           
        }
        return jsonify(data)


@app.route('/dayScheudle', methods=['GET', "POST"])
def get_data_day():
    if request.method == "POST":
        wdadawdsdasdasdsac here 
    return 0
    wb = wbFull['Table 1']
    wb1 = wbFull['Table 2']

    group = group

    limit = 200
    baseRow = 2
    row = 2
    column = 1
    ret = []

    while True:
        valueOfCell = wb.cell(row=row, column=column).value
        if valueOfCell == None and row == limit and column > 10:
            break
        elif valueOfCell == None and row == limit:
            row = baseRow
            column += 1
        else:                                     
            if valueOfCell == group or group in str(valueOfCell):
                q1 = str(wb.cell(row=row, column=column-1).value)
                ret.append(q1)
                row = baseRow
                column += 1
            else:
                row += 1
    if True:
        limit = 200
        baseRow = 2
        row = 2
        column = 1
        while True:
            valueOfCell = wb1.cell(row=row, column=column).value
            if valueOfCell == None and row == limit and column > 10:
                break
            elif valueOfCell == None and row == limit:
                row = baseRow
                column += 1
            else:                       
                if valueOfCell == group or group in str(valueOfCell):
                    q1 = str(wb1.cell(row=row, column=column-1).value)
                    ret.append(q1)
                    row = baseRow
                    column += 1
                else:
                    row += 1

def getscheudle(group, date):
    wbFull = load_workbook(date)
    wb = wbFull['Table 1']
    wb1 = wbFull['Table 2']

    group = group

    limit = 200
    baseRow = 2
    row = 2
    column = 1
    ret = []

    while True:
        valueOfCell = wb.cell(row=row, column=column).value
        if valueOfCell == None and row == limit and column > 10:
            break
        elif valueOfCell == None and row == limit:
            row = baseRow
            column += 1
        else:                                     
            if valueOfCell == group or group in str(valueOfCell):
                q1 = str(wb.cell(row=row, column=column-1).value)
                ret.append(q1)
                row = baseRow
                column += 1
            else:
                row += 1
    if True:
        limit = 200
        baseRow = 2
        row = 2
        column = 1
        while True:
            valueOfCell = wb1.cell(row=row, column=column).value
            if valueOfCell == None and row == limit and column > 10:
                break
            elif valueOfCell == None and row == limit:
                row = baseRow
                column += 1
            else:                       
                if valueOfCell == group or group in str(valueOfCell):
                    q1 = str(wb1.cell(row=row, column=column-1).value)
                    ret.append(q1)
                    row = baseRow
                    column += 1
                else:
                    row += 1
    return ret


if __name__ == '__main__':
    app.run(debug=True, port=5000)